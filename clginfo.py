import openpyxl

wb = openpyxl.load_workbook('ccode.xlsx')
sh = wb.active

#c1 = sh['A2']
#Using cell() function
#col=2,3,4
#row=3-204
c3 = sh.cell(row=3,column=2)
d={}
for i in range(3,205,1):
    e={}
    e['C_name']=sh.cell(row=i,column=3).value
    e['City']=sh.cell(row=i,column=4).value
    e['Code']=sh.cell(row=i,column=2).value
    d[sh.cell(row=i,column=2).value]=e
    #d[sh.cell(row=i,column=2).value]={sh.cell(row=i,column=3).value,sh.cell(row=i,column=4).value}
    
    
    
#print(d)
def getclg(code):
    try:
        return d[code]
    except:
        return 'NON JNTUH'
#print(getclg('VE'))